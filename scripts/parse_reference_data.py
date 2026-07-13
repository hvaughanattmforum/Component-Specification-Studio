"""
Conversion of the official TMForum GB921 (eTOM), GB922 (SID), and GB1033/
GB1033F (Functional Framework) Excel exports into JSON catalogs the wizard's
Node server can serve directly, without adding an xlsx dependency to Node.

This script lives in the app's own code (component-spec-editor/scripts/),
separate from the frameworks data directory it operates on - that directory
should only ever contain the source spreadsheets and the generated JSON, no
code. Pass the frameworks directory as the first argument; if omitted, the
current working directory is used (so `cd frameworks && python
<path-to>/parse_reference_data.py` still works the same as before).

Supports multiple versions side by side: every GB921*.xlsx / GB922*.xlsx /
GB1033*.xlsx file found in that directory is converted independently, and
the version parsed from its filename (e.g. "..._v26.0.xlsx" -> "v26.0") is
baked into the output filename (etom_v26.0.json, sid_v26.0.json, ...) and
the "version" field inside it. Drop in a new release's spreadsheet - e.g.
GB921_..._v27.0.xlsx - re-run this script, and both v26.0 and v27.0 catalogs
exist side by side; the server picks the latest by default.

If a file's version can't be parsed from its name, this does not fail loudly:
it falls back to "_" (a literal underscore) in place of the version segment
in the output filename, e.g. "etom__.json", so the file is still produced
and clearly flagged as unversioned rather than silently overwriting/crashing.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

FRAMEWORKS_DIR = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else Path.cwd()


def token(text):
    return re.sub(r"\s+", "_", str(text).strip())


# Matches "..._v26.0.xlsx" / "..._V27.xlsx" etc. at the end of a filename.
VERSION_RE = re.compile(r"_[vV]([0-9]+(?:\.[0-9]+)*)\.xlsx$")


def extract_version(filename):
    m = VERSION_RE.search(filename)
    return f"v{m.group(1)}" if m else None


def version_slug(version):
    """Filename-safe stand-in for a version - literal underscore(s) when unknown."""
    return version if version else "_"


def find_files(prefix):
    return sorted(FRAMEWORKS_DIR.glob(f"{prefix}*.xlsx"))


def find_etom_sheet(wb):
    # Sheet name is version-suffixed (e.g. "eTOM26.0"), so match by prefix
    # rather than hardcoding one version's exact name.
    candidates = [s for s in wb.sheetnames if s.lower().startswith("etom") and "delete" not in s.lower()]
    return wb[candidates[0] if candidates else wb.sheetnames[0]]


def parse_etom(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = find_etom_sheet(wb)
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            continue  # sparse/trailing row shorter than the sheet's real columns
        name, identifier, level, _ext, _brief, domain = row[0], row[1], row[2], row[3], row[4], row[5]
        if not name or not identifier:
            continue
        entries.append({
            "id": str(identifier).strip(),
            "name": str(name).strip(),
            "token": token(name),
            "domain": domain,
            "level": level,
        })
    return entries


def find_functions_sheet(wb):
    # Sheet name isn't consistent across releases - usually exactly
    # "Functions", but some versions only have "Functions " (trailing space)
    # or "Functions and AFs" instead. Prefer an exact (whitespace/case
    # insensitive) match; fall back to any sheet name starting with
    # "functions" so those variants are still found rather than silently
    # falling through to the title-page sheet and producing zero entries.
    normalized = {s.strip().lower(): s for s in wb.sheetnames}
    if "functions" in normalized:
        return wb[normalized["functions"]]
    candidates = [s for s in wb.sheetnames if s.strip().lower().startswith("functions")]
    return wb[candidates[0]] if candidates else wb[wb.sheetnames[0]]


def _header_key(value):
    """Normalizes a header cell for matching - e.g. "AF Lev.2'" and "AF Lev.2 " both become "aflev2"."""
    return re.sub(r"[^a-z0-9]", "", str(value).lower()) if value is not None else ""


def parse_functional_framework(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = find_functions_sheet(wb)
    rows = ws.iter_rows(values_only=True)

    # Column order isn't consistent across releases - e.g. "Vertical" sits
    # right after "Domain" in some versions and at the very end in others -
    # so columns are looked up by their header name instead of a fixed
    # position, which would silently read the wrong column depending on
    # which release's layout happened to be used.
    header = next(rows, None)
    if not header:
        return []
    columns = {}
    for i, cell in enumerate(header):
        key = _header_key(cell)
        if key:
            columns.setdefault(key, i)

    def get(row, key):
        i = columns.get(key)
        return row[i] if i is not None and i < len(row) else None

    entries = []
    for row in rows:
        name = get(row, "functionname")
        func_id = get(row, "functionid")
        if not name or func_id is None:
            continue
        if "(deleted)" in str(name).lower():
            continue
        entries.append({
            "id": str(func_id).strip(),
            "name": str(name).strip(),
            "token": token(name),
            "domain": get(row, "domain"),
            "vertical": get(row, "vertical"),
            "af1": get(row, "aflev1"),
            "af2": get(row, "aflev2"),
        })
    return entries


SID_DOMAIN_SHEETS = [
    "Market", "Sales", "Customer", "Product", "Service", "Resource",
    "Business Partner", "Enterprise", "Patterns", "Shared",
]


def parse_sid(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    pairs = {}
    triples = {}
    # Each sheet is one fixed domain. Within it, ABEs can nest (e.g. a row
    # "Product Configuration ABE.Product Action ABE" is a *sub*-ABE, not a new
    # domain), so only accept a col-A value as a top-level ABE when its prefix
    # is exactly this sheet's domain name; anything else is a nested header we
    # skip without losing the current (domain, abe) context for the BE rows
    # that follow it.
    for sheet_name in SID_DOMAIN_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        domain = f"{sheet_name} Domain"
        prefix = f"{domain}."
        abe = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 2:
                continue  # sparse/trailing row shorter than the sheet's real columns
            a, b = row[0], row[1]
            if a is not None:
                a = str(a).strip()
                if a.startswith(prefix):
                    abe = a[len(prefix):].strip()
                    pairs.setdefault(domain, set()).add(abe)
                elif abe and a.startswith(f"{abe}."):
                    # A nested sub-ABE (e.g. "Loyalty ABE.Loyalty Program ABE"):
                    # real specs sometimes reference these as a third SID
                    # segment alongside plain BEs, so offer both as "children".
                    child = a[len(abe) + 1:].strip()
                    triples.setdefault((domain, abe), set()).add(("ABE", child))
                # else: unrecognized nesting - ignore, keep current abe context
            elif b is not None and abe:
                b = str(b).strip()
                triples.setdefault((domain, abe), set()).add(("BE", b))

    domains = sorted(pairs.keys())
    abes_by_domain = {d: sorted(abes) for d, abes in pairs.items()}
    bes_by_domain_abe = {
        f"{d}||{a}": sorted([{"kind": kind, "name": name} for kind, name in children], key=lambda c: c["name"])
        for (d, a), children in triples.items()
    }
    return {
        "domains": domains,
        "abesByDomain": abes_by_domain,
        "besByDomainAbe": bes_by_domain_abe,
        "domainToken": {d: token(d) for d in domains},
        "abeToken": {a: token(a) for abes in abes_by_domain.values() for a in abes},
    }


CONVERTERS = [
    ("GB921", "etom", parse_etom, lambda entries: {"entries": entries}),
    # "GB1033" (not just "GB1033F") so older exports named without the "F" -
    # e.g. GB1033_Functional_Framework_Excel_v23.0.0.xlsx - are picked up too.
    ("GB1033", "functionalFramework", parse_functional_framework, lambda entries: {"entries": entries}),
    ("GB922", "sid", parse_sid, lambda payload: payload),
]


def main():
    written = []
    for prefix, out_name, parser, shape in CONVERTERS:
        files = find_files(prefix)
        if not files:
            print(f"No {prefix}*.xlsx found - skipping {out_name}")
            continue
        for xlsx_path in files:
            version = extract_version(xlsx_path.name)
            if version is None:
                print(f"Warning: could not parse a version from '{xlsx_path.name}' - "
                      f"writing {out_name}_{version_slug(version)}.json with version left unset")
            result = parser(xlsx_path)
            payload = {"version": version, **shape(result)}
            out_path = FRAMEWORKS_DIR / f"{out_name}_{version_slug(version)}.json"
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False)
            written.append(out_path.name)

    print("Wrote:", ", ".join(written) if written else "(nothing)")


if __name__ == "__main__":
    main()
